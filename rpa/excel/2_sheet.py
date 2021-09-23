# pip install openpyxl --trusted-host pypi.python.org --trusted-host files.pythonhosted.org --trusted-host pypi.org

from openpyxl import Workbook
wb = Workbook() #새 워크북 생성 
ws = wb.create_sheet()
ws.title = "mySheet"
ws.sheet_properties.tabColor = "ff66ff" # RGB 필터로 색을 넣어줌 

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 0)

# 딕셔너리 형태로 sheet에 접근 
new_ws = wb["NewSheet"]
print(wb.sheetnames)


# sheet 복사 
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")

