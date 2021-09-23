from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active 

# 2열에 새로운 열추가 
# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "국어"

# 번호 영어 수학
ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")