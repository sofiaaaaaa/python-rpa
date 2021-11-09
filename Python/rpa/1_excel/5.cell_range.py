from openpyxl import Workbook
from random import * 
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

ws.append(["ㅁ", "ㅂ", "ㅅ"])

for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])


col_B = ws["B"] # ㅂ 컬럼만 가지고오기
print(col_B)

for cell in col_B:
    print(cell.value)


col_range = ws["B:C"] # ㅂ, ㅅ 컬럼 가져오기 

for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가져오기

for cell in row_title:
    print(cell.value)

row_raneg = ws[2:6] # 두번째 줄에서 6번재 줄까지 가져오기 
row_raneg = ws[2:ws.max_row] # 두번째 줄에서 마지막 줄까지 가져오기 

for rows in row_raneg:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")
            
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

# 전체 rows
print(tuple(ws.rows))

for row in tuple(ws.rows):
    print(row[1].value)

# 전체 columns
print(tuple(ws.columns))

for column in tuple(ws.columns):
    print(column[1].value)



for row in ws.iter_rows():
    print(row[1].value)

for col in ws.iter_cols():
    print(col[0].value)

#  1번째 줄에서 5번재 줄까지, 2번째열에서 3번째 열까지 
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)

wb.save("sample.xlsx")



